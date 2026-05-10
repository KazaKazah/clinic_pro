from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from outpatient.models import ICD10Diagnosis


class Command(BaseCommand):
    help = "Импортирует справочник МКБ-10 из Excel файла"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Путь к Excel файлу .xlsx")
        parser.add_argument(
            "--sheet",
            type=str,
            default="МКБ-10",
            help="Имя листа Excel (по умолчанию: МКБ-10)",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Удалить существующие записи перед импортом",
        )
        parser.add_argument(
            "--source-url",
            type=str,
            default="https://adilet.zan.kz/rus/docs/V2000021784",
            help="Ссылка на источник",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        sheet_name = options["sheet"]
        replace = options["replace"]
        source_url = options["source_url"]

        if not excel_path.exists():
            raise CommandError(f"Файл не найден: {excel_path}")

        wb = load_workbook(excel_path, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise CommandError(
                f"Лист '{sheet_name}' не найден. Доступные листы: {', '.join(wb.sheetnames)}"
            )

        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)

        if not header:
            raise CommandError("Excel-файл пустой.")

        normalized_header = [str(h).strip().lower() if h is not None else "" for h in header]

        try:
            code_index = normalized_header.index("code")
            name_index = normalized_header.index("name")
        except ValueError:
            raise CommandError(
                "Не найдены обязательные колонки 'Code' и 'Name'."
            )

        if replace:
            deleted_count, _ = ICD10Diagnosis.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"Удалено записей: {deleted_count}"))

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in rows:
            if not row:
                continue

            code = row[code_index] if code_index < len(row) else None
            name = row[name_index] if name_index < len(row) else None

            if code is None or name is None:
                skipped_count += 1
                continue

            code = str(code).strip()
            name = str(name).strip()

            if not code or not name:
                skipped_count += 1
                continue

            obj, created = ICD10Diagnosis.objects.update_or_create(
                code=code,
                defaults={
                    "name": name,
                    "is_active": True,
                    "external_url": source_url,
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(self.style.SUCCESS("Импорт завершен."))
        self.stdout.write(f"Создано: {created_count}")
        self.stdout.write(f"Обновлено: {updated_count}")
        self.stdout.write(f"Пропущено: {skipped_count}")
